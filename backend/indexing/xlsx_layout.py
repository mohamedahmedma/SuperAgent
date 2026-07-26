"""XLSX layout parsing to the shared block contract (heading/text/table).

Each worksheet becomes a level-1 heading block (the sheet name is the natural
section/topic) followed by one table block with the sheet's populated rows — so
spreadsheet content lands in the index as real, row-safe table chunks under their
sheet's topic instead of the flattened prose the UnstructuredExcelLoader produced.

The sheet heading between consecutive sheets also guarantees the cross-page table
stitcher never merges two different sheets' tables (they are never adjacent).

page_number is the worksheet index (the citation anchor: "sheet 2"); `top` is a
document-order surrogate. Sheets are row-capped defensively so a pathological
workbook cannot balloon ingestion. openpyxl is imported lazily.
"""
from __future__ import annotations

import logging
from typing import Any, Dict, List

from backend.indexing.pdf_layout import format_table_rows, normalize_table_rows

logger = logging.getLogger(__name__)

# Defensive ceiling per sheet; rows beyond it are dropped with a warning.
MAX_ROWS_PER_SHEET = 5000


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def parse_xlsx_blocks(file_path: str) -> List[Dict[str, Any]]:
    """Parse a .xlsx into ordered heading/table blocks. Raises on unreadable
    files — the caller decides whether to fall back to flat extraction."""
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        blocks: List[Dict[str, Any]] = []
        order = 0.0
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            raw_rows: List[List[str]] = []
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if row_index >= MAX_ROWS_PER_SHEET:
                    logger.warning(
                        "Sheet %r exceeds %d rows — truncating remaining rows",
                        worksheet.title, MAX_ROWS_PER_SHEET,
                    )
                    break
                raw_rows.append([_cell_to_text(value) for value in row])

            rows = normalize_table_rows(raw_rows)
            if not rows:
                continue

            title = (worksheet.title or "").strip()
            if title:
                blocks.append({
                    "type": "heading",
                    "content": title,
                    "level": 1,
                    "page_number": sheet_index,
                    "top": order,
                })
                order += 1.0
            blocks.append({
                "type": "table",
                "content": format_table_rows(rows),
                "rows": rows,
                "page_number": sheet_index,
                "top": order,
            })
            order += 1.0
        return blocks
    finally:
        workbook.close()
